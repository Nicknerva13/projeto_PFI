from django.views.decorators.csrf import csrf_exempt
from .forms import LoginForm, ImagemCarrosselForm, UsuarioForm
from .models import Usuario, ImagemCarrossel, Produto, Venda, ItemVenda, HistoricoCompra


# ========================================================
# 1. PÁGINAS PRINCIPAIS E AUTENTICAÇÃO
# ========================================================

def home(request):
    imagens = ImagemCarrossel.objects.filter(ativo=True)
    return render(request, 'Site_principal/tela_inicial/index.html', {'imagens': imagens})

def historico_adm(request):
    # Verifica se é admin
    if not request.session.get("is_admin"):
        return redirect("login")

    # Busca todas as vendas do sistema
    historico = Venda.objects.all().order_by("-data")

    return render(request, "Site_principal/Historico/historico_adm.html", {
        "historico": historico
    })


def login_pagina(request):
    """Renderiza apenas a tela de ‘login’ (se acessada via GET direto)"""
    return render(request, 'Site_principal/tela_login/index.html')

def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()

        # Tenta achar usuário por email ou matrícula
        usuario = Usuario.objects.filter(
            Q(email__iexact=username) | Q(matricula=username)
        ).first()

        if not usuario:
            messages.error(request, "Usuário não encontrado!")
            return render(request, "Site_principal/tela_login/index.html")

        # Valida senha
        if usuario.senha != password:
            messages.error(request, "Senha incorreta!")
            return render(request, "Site_principal/tela_login/index.html")

        # ------------------------
        # LOGIN BEM-SUCEDIDO
        # ------------------------
        request.session["usuario_id"] = usuario.id
        request.session["usuario_nome"] = usuario.nome
        request.session["is_admin"] = usuario.is_admin

        # Redireciona corretamente
        return redirect("adm_dashboard" if usuario.is_admin else "cliente")

    return render(request, "Site_principal/tela_login/index.html")

def logout_view(request):
    request.session.flush()
    return redirect("home")

def registrar(request):
    if request.method == "POST":
        nome = request.POST.get("nome")
        matricula = request.POST.get("matricula")
        email = request.POST.get("email")
        senha = request.POST.get("senha")
        senhaConfirm = request.POST.get("senhaConfirm")

        telefone = request.POST.get("telefone")  # <-- ADICIONADO
        telefone_responsavel = request.POST.get("telefone_responsavel")  # <-- ADICIONADO

        if senha != senhaConfirm:
            messages.error(request, "As senhas não conferem!")
            return redirect("registrar")

        if Usuario.objects.filter(email=email).exists():
            messages.error(request, "Email já cadastrado!")
            return redirect("registrar")

        if Usuario.objects.filter(matricula=matricula).exists():
            messages.error(request, "Matrícula já cadastrada!")
            return redirect("registrar")

        Usuario.objects.create(
            nome=nome,
            matricula=matricula,
            email=email,
            senha=senha,
            telefone=telefone,  # <-- SALVA
            telefone_responsavel=telefone_responsavel  # <-- SALVA
        )

        messages.success(request, "Cadastro realizado com sucesso!")
        return redirect("login")

    return render(request, "Site_principal/tela_registrar/index.html")
def cliente(request):
    if not request.session.get("usuario_id"):
        return redirect("login")
    usuario = Usuario.objects.get(id=request.session["usuario_id"])
    return render(request, "Site_principal/tela_home_cliente/index.html", {"usuario": usuario})

def adm_dashboard(request):
    if not request.session.get("usuario_id") or not request.session.get("is_admin"):
        return redirect("login")
    usuario = Usuario.objects.get(id=request.session["usuario_id"])
    return render(request, "Site_principal/tela_home_adm/index.html", {"usuario": usuario})

# ========================================================
# 2. PAINÉIS (ADM e CLIENTE)
# ========================================================

def pagina_adm(request):
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return redirect("login")

    admin_user = get_object_or_404(Usuario, id=usuario_id)
    if not admin_user.is_admin:
        return redirect("cliente")

    todos_usuarios = Usuario.objects.all()

    dados = []

    for u in todos_usuarios:
        total_devido = Venda.objects.filter(usuario=u, pago=False)\
            .aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal('0')

        # ✔ comparação correta com Decimal
        if total_devido > Decimal('0'):
            dados.append({
                'usuario': u,
                'total_dividas': total_devido
            })
    print("DADOS ENVIADOS:", dados)
    return render(request, "Site_principal/tela_home_adm/index.html", {
        "usuario": admin_user,
        "dados": dados
    })

def login_usuario(request):
    if request.method == 'POST':
        matricula = request.POST.get('matricula')
        senha = request.POST.get('senha')

        try:
            usuario = Usuario.objects.get(matricula=matricula, senha=senha)  # senha simples para teste
            request.session['matricula'] = usuario.matricula
            return redirect('/cliente/')
        except Usuario.DoesNotExist:
            return render(request, 'Site_principal/tela_home_cliente/index.html', {'erro': 'Usuário ou senha inválidos'})
    return render(request, 'Site_principal/tela_home_cliente/index.html')

from django.contrib.auth.hashers import make_password

# =========================================================
#  EDITAR MEUS DADOS (ADMIN)
# =========================================================
def editar_meus_dados_adm(request):
    usuario_id = request.session.get("usuario_id")

    # Usuário não logado
    if not usuario_id:
        return redirect("login")

    usuario = get_object_or_404(Usuario, id=usuario_id)

    # ❗ Impede usuário comum de acessar página ADM
    if not usuario.is_admin:
        messages.error(request, "Você não tem permissão para acessar esta página.")
        return redirect("cliente")

    if request.method == "POST":
        # Atualização básica
        usuario.nome = request.POST.get("nome", usuario.nome)
        usuario.email = request.POST.get("email", usuario.email)
        usuario.telefone = request.POST.get("telefone", usuario.telefone)
        usuario.telefone_responsavel = request.POST.get("telefone_responsavel", usuario.telefone_responsavel)

        # Senha
        nova_senha = request.POST.get("senha")
        if nova_senha:
            usuario.senha = make_password(nova_senha)

        # Foto
        foto = request.FILES.get("foto")
        if foto:
            usuario.foto = foto

        usuario.save()
        messages.success(request, "Dados atualizados com sucesso!")
        return redirect("adm_dashboard")

    return render(
        request,
        "Site_principal/Editar_meus_dados/editar_meus_dados_adm.html",
        {"usuario": usuario}
    )

# =========================================================
#  EDITAR MEUS DADOS (USUÁRIO COMUM)
# =========================================================
def editar_meus_dados_usuario(request):
    usuario_id = request.session.get("usuario_id")

    # Usuário não logado
    if not usuario_id:
        return redirect("login")

    usuario = get_object_or_404(Usuario, id=usuario_id)

    # ❗ Impede admin de acessar tela de usuário
    if usuario.is_admin:
        messages.error(request, "Admins não editam dados por esta página.")
        return redirect("cliente")

    if request.method == "POST":
        # Atualização básica
        usuario.nome = request.POST.get("nome", usuario.nome)
        usuario.email = request.POST.get("email", usuario.email)
        usuario.telefone = request.POST.get("telefone", usuario.telefone)
        usuario.telefone_responsavel = request.POST.get("telefone_responsavel", usuario.telefone_responsavel)

        # Senha
        nova_senha = request.POST.get("senha")
        if nova_senha:
            usuario.senha = make_password(nova_senha)

        # Foto
        foto = request.FILES.get("foto")
        if foto:
            usuario.foto = foto

        usuario.save()
        messages.success(request, "Dados atualizados com sucesso!")
        return redirect("cliente")

    return render(
        request,
        "Site_principal/Editar_meus_dados/editar_meus_dados_usuario.html",
        {"usuario": usuario}
    )

# ========================================================
# 3. GERENCIAMENTO DE PRODUTOS
# ========================================================

def listar_produtos(request):
    produtos = Produto.objects.all()
    return render(request, 'Site_principal/editar_produto/listar_produto.html', {'produtos': produtos})

def adicionar_produto(request):
    if request.method == "POST":
        nome = request.POST.get("nome")
        preco_compra = Decimal(request.POST.get("preco_compra"))
        preco_venda = Decimal(request.POST.get("preco_venda"))
        estoque = int(request.POST.get("estoque"))
        foto = request.FILES.get("foto")

        produto = Produto.objects.create(
            nome=nome,
            preco_compra=preco_compra,
            preco_venda=preco_venda,
            estoque=estoque,
            foto=foto
        )

        # 🔥 REGISTRAR A ENTRADA INICIAL DE ESTOQUE
        EntradaEstoque.objects.create(
            produto=produto,
            quantidade=estoque,
            preco_unitario=preco_compra
        )

        return redirect("listar_produtos")

    return render(request, "Site_principal/editar_produto/adicionar_produto.html")

from django.shortcuts import render, get_object_or_404, redirect
def editar_produto(request, id):
    produto = get_object_or_404(Produto, id=id)

    if request.method == "POST":
        produto.nome = request.POST.get("nome")
        produto.preco_compra = request.POST.get("preco_compra")
        produto.preco_venda = request.POST.get("preco_venda")
        produto.estoque = request.POST.get("estoque")

        if 'foto' in request.FILES:
            produto.foto = request.FILES['foto']

        produto.save()
        return redirect('listar_produtos')

    return render(request, "Site_principal/editar_produto/editar_produto.html", {
        "produto": produto
    })

def excluir_produto(request, produto_id):
    produto = get_object_or_404(Produto, id=produto_id)
    produto.delete()
    return redirect('listar_produtos')
def adicionar_estoque(request, produto_id):
    produto = get_object_or_404(Produto, id=produto_id)

    if request.method == 'POST':
        qtd = int(request.POST.get('quantidade', 0))

        # Atualiza o estoque
        produto.estoque += qtd
        produto.save()

        # 🔥 REGISTRAR A ENTRADA DE ESTOQUE
        EntradaEstoque.objects.create(
            produto=produto,
            quantidade=qtd,
            preco_unitario=produto.preco_compra
        )

        messages.success(request, f"Estoque atualizado! Novo total: {produto.estoque}")
        return redirect('listar_produtos')

    return render(request, 'Site_principal/editar_produto/adicionar_estoque.html', {'produto': produto})



# ========================================================
# 4. VENDAS E API (Para o JavaScript)
# ========================================================

def vendas_view(request):
    """Renderiza a tela de vendas"""
    produtos = Produto.objects.all()
    return render(request, 'Site_principal/tela_vendas/index.html', {'produtos': produtos})


def api_listar_produtos(request):
    """API para preencher selects via JS se necessário"""
    produtos = Produto.objects.all()
    data = [{'id': p.id, 'nome': p.nome, 'preco': float(p.preco)} for p in produtos]
    return JsonResponse(data, safe=False)

from django.db.models import Q

def api_buscar_usuario(request):
    """
    API para buscar usuários por matrícula, por nome (busca parcial) ou por um termo genérico.
    Retorna lista de {id, nome, matricula, saldo_em_dividas}.
    """
    matricula = request.GET.get('matricula')
    nome = request.GET.get('nome')
    buscar = request.GET.get('buscar')
    qs = Usuario.objects.none()

    if matricula:
        qs = Usuario.objects.filter(matricula__iexact=matricula)
    elif nome:
        qs = Usuario.objects.filter(nome__icontains=nome)
    elif buscar:
        qs = Usuario.objects.filter(Q(matricula__iexact=buscar) | Q(nome__icontains=buscar))
    else:
        qs = Usuario.objects.none()

    dados = []
    for u in qs:
        dados.append({
            'id': u.id,
            'nome': u.nome,
            'matricula': u.matricula,
            'saldo_em_dividas': float(u.saldo_em_dividas or 0),
        })

    return JsonResponse(dados, safe=False)

from django.http import JsonResponse
import json
from .models import Usuario, Produto, Venda, ItemVenda

def api_registrar_venda(request):
    if request.method != "POST":
        return JsonResponse({"sucesso": False, "erro": "Método inválido"})

    try:
        data = json.loads(request.body)
        matricula = data.get("matricula")
        produtos_data = data.get("produtos", [])
        tipo_pagamento = data.get("metodo_pagamento", "fiado")

        # CLIENTE da compra
        usuario = Usuario.objects.get(matricula=matricula)
        pago = tipo_pagamento != "fiado"

        # VENDEDOR (usuário logado)
        vendedor = Usuario.objects.get(email=request.user.email)

        nova_venda = Venda.objects.create(
            usuario=usuario,
            vendedor=vendedor,
            valor_total=0,
            pago=pago,
            metodo_pagamento=tipo_pagamento
        )

        total_venda = 0

        for item in produtos_data:
            prod_id = item.get("id")
            quantidade = int(item.get("quantidade", 1))
            produto = Produto.objects.get(id=prod_id)
            preco_unitario = produto.preco_venda
            subtotal = preco_unitario * quantidade

            ItemVenda.objects.create(
                venda=nova_venda,
                produto=produto,
                quantidade=quantidade,
                preco_unitario=preco_unitario
            )

            total_venda += subtotal

            if produto.estoque >= quantidade:
                produto.estoque -= quantidade
                produto.save()

        nova_venda.valor_total = total_venda
        nova_venda.save()

        if tipo_pagamento == "fiado":
            usuario.saldo_em_dividas += total_venda
            if not usuario.data_primeira_compra:
                usuario.data_primeira_compra = timezone.now().date()
            usuario.save()

        return JsonResponse({"sucesso": True, "total_venda": float(total_venda)})

    except Usuario.DoesNotExist:
        return JsonResponse({"sucesso": False, "erro": "Usuário não encontrado"})
    except Produto.DoesNotExist:
        return JsonResponse({"sucesso": False, "erro": "Produto não encontrado"})
    except Exception as e:
        return JsonResponse({"sucesso": False, "erro": str(e)})

# ========================================================
# 5. GESTÃO DE DÍVIDAS E PAGAMENTOS
# ========================================================

def minhas_dividas(request):
    """Mostra as dívidas de TODOS os usuários para o administrador"""

    usuario_id = request.session.get('usuario_id')
    usuario = get_object_or_404(Usuario, id=usuario_id)

    # Garante que só o ADM vê isso
    if not usuario.is_admin:
        return redirect("cliente")

    # Monta lista com total devido por cada usuário
    usuarios = Usuario.objects.all()
    dados = []

    for u in usuarios:
        total = Venda.objects.filter(usuario=u, pago=False).aggregate(
            total=Sum('valor_total')
        )['total'] or Decimal('0')

        dados.append({
            'usuario': u,
            'total_devido': total
        })

    return render(
        request,
        'Site_principal/dividas/minhas_dividas_adm.html',
        {'dados': dados, 'usuario': usuario}
    )

def minhas_dividas_usuario(request):
    usuario_id = request.session.get("usuario_id")

    if not usuario_id:
        messages.error(request, "Você precisa estar logado para acessar suas dívidas.")
        return redirect("login")

    # Usuário logado
    usuario = Usuario.objects.get(id=usuario_id)

    # Vendas NÃO pagas desse usuário
    vendas = (
        Venda.objects.filter(usuario=usuario, pago=False)
        .order_by("-data")   # campo 'data' existe no seu modelo
    )

    # Total devido
    total_devido = vendas.aggregate(Sum("valor_total"))["valor_total__sum"] or Decimal("0.00")

    # 🔹 Limite fixo
    limite = Decimal("30.00")
    limite_restante = limite - total_devido

    # Montar estrutura de detalhes para o template
    detalhes = []
    for venda in vendas:
        itens = []
        for item in venda.itens.all():
            itens.append({
                "produto": item.produto.nome,
                "quantidade": item.quantidade,
                "preco": item.preco_unitario,
                "total_item": item.preco_unitario * item.quantidade,
            })

        detalhes.append({
            "id": venda.id,
            "data": venda.data,
            "valor_total": venda.valor_total,
            "itens": itens,
        })

    return render(request, "Site_principal/dividas/minhas_dividas_usuario.html", {
        "usuario": usuario,
        "detalhes": detalhes,
        "total_devido": total_devido,
        "limite": limite,
        "limite_restante": limite_restante,
    })

def registrar_pagamento(request, matricula):
    """ADM registra pagamento e baixa as dívidas mais antigas"""
    usuario = get_object_or_404(Usuario, matricula=matricula)

    # Calcula total devido
    vendas_abertas = Venda.objects.filter(usuario=usuario, pago=False).order_by('data')
    total_devido = vendas_abertas.aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal(0)

    if request.method == 'POST':
        valor_pago = Decimal(request.POST.get('valor_pago', 0))
        restante = valor_pago

        # Algoritmo: quitar dívidas da mais antiga para a mais nova
        for venda in vendas_abertas:
            if restante <= 0:
                break

            if restante >= venda.valor_total:
                # Paga a venda inteira
                restante -= venda.valor_total
                venda.pago = True
                venda.save()
            else:
                break
        novo_total = Venda.objects.filter(usuario=usuario, pago=False).aggregate(
            Sum('valor_total')
        )['valor_total__sum'] or Decimal(0)

        usuario.saldo_em_dividas = novo_total
        usuario.save()

        messages.success(request, f"Pagamento de R$ {valor_pago} registrado!")
        return redirect('adm_dashboard')

    return render(request, 'Site_principal/dividas/registrar_pagamento.html', {
        'usuario_alvo': usuario,
        'total_devido': total_devido
    })



# ========================================================
# 6. OUTROS (Carrossel, Relatórios, Usuários)
# ========================================================

def listar_usuarios(request):
    usuarios = Usuario.objects.all()
    return render(request, "Site_principal/editar_usuario/gerenciar_usuarios.html", {"usuarios": usuarios})
from .forms import UsuarioForm
from django.contrib import messages

def adicionar_usuario(request):
    if request.method == "POST":
        nome = request.POST.get("nome")
        matricula = request.POST.get("matricula")
        telefone = request.POST.get("telefone")
        telefone_responsavel = request.POST.get("telefone_responsavel")
        email = request.POST.get("email")
        senha = request.POST.get("senha")
        is_admin = request.POST.get("is_admin") == "on"

        # Validar email único
        if Usuario.objects.filter(email=email).exists():
            messages.error(request, "Email já cadastrado!")
            return redirect("adicionar_usuario")

        # Validar matrícula única
        if Usuario.objects.filter(matricula=matricula).exists():
            messages.error(request, "Matrícula já cadastrada!")
            return redirect("adicionar_usuario")

        # Salvar usuário (sem foto)
        Usuario.objects.create(
            nome=nome,
            matricula=matricula,
            telefone=telefone,
            telefone_responsavel=telefone_responsavel,
            email=email,
            senha=senha,    # SEM CRIPTOGRAFAR
            is_admin=is_admin,
        )

        messages.success(request, "Usuário adicionado com sucesso!")
        return redirect("listar_usuarios")

    return render(request, "Site_principal/editar_usuario/adicionar_usuario.html")



def editar_usuario(request, id):
    usuario = get_object_or_404(Usuario, id=id)
    if request.method == "POST":
        form = UsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            return redirect("listar_usuarios")
    else:
        form = UsuarioForm(instance=usuario)
    return render(request, "Site_principal/editar_usuario/editar_usuario.html", {"form": form, "usuario": usuario})


def excluir_usuario(request, id):
    usuario = get_object_or_404(Usuario, id=id)
    usuario.delete()
    return redirect("listar_usuarios")


# Carrossel
def editar_carrossel(request):
    imagens = ImagemCarrossel.objects.all()
    return render(request, 'Site_principal/editar_carrosel/index.html', {'imagens': imagens})


def adicionar_imagem(request):
    if request.method == 'POST':
        form = ImagemCarrosselForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('editar_carrossel')
    else:
        form = ImagemCarrosselForm()
    return render(request, 'Site_principal/editar_carrosel/adicionar_imagem.html', {'form': form})


def remover_imagem(request, imagem_id):
    imagem = get_object_or_404(ImagemCarrossel, id=imagem_id)
    imagem.delete()
    return redirect('editar_carrossel')


def editar_imagem(request, imagem_id):
    imagem = get_object_or_404(ImagemCarrossel, id=imagem_id)
    if request.method == 'POST':
        titulo = request.POST.get('titulo', imagem.titulo)
        if 'imagem' in request.FILES:
            imagem.url = request.FILES['imagem']
        imagem.titulo = titulo
        imagem.save()
        return redirect('editar_carrossel')
    return render(request, 'Site_principal/editar_carrosel/editar_imagem.html', {'imagem': imagem})


@csrf_exempt
def ativar_imagem(request, id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            imagem = ImagemCarrossel.objects.get(id=id)
            imagem.ativo = data.get('ativo', False)
            imagem.save()
            return JsonResponse({'success': True})
        except:
            return JsonResponse({'success': False})
    return JsonResponse({'success': False})

from django.shortcuts import render, redirect

def historico_usuario(request):
    # Verifica se o usuário está logado
    usuario_id = request.session.get("usuario_id")
    if not usuario_id:
        return redirect("login")

    # Busca o usuário e o histórico de vendas
    usuario = Usuario.objects.get(id=usuario_id)
    historico = Venda.objects.filter(usuario=usuario).order_by('-data')[:10]
    return render(request, "Site_principal/Historico/historico_usuario.html", {"historico": historico, "usuario": usuario})

from django.shortcuts import render
from django.utils import timezone

def gerar_relatorio(request):
    periodo = request.GET.get("periodo", "hoje")

    hoje = timezone.now().astimezone().date()

    periodos = {
        "hoje": hoje,
        "5d": hoje - timedelta(days=5),
        "10d": hoje - timedelta(days=10),
        "1m": hoje - timedelta(days=30),
        "6m": hoje - timedelta(days=180),
        "1a": hoje - timedelta(days=365),
        "2a": hoje - timedelta(days=730),
        "3a": hoje - timedelta(days=1095),
    }

    data_inicial = periodos.get(periodo, hoje)
    vendas = Venda.objects.filter(data__date__gte=data_inicial) \
                          .select_related('usuario')
    resumo_clientes = {}
    total_fiado = 0
    total_pago = 0
    total_vendas = 0

    for venda in vendas:
        uid = venda.usuario.id
        if uid not in resumo_clientes:
            resumo_clientes[uid] = {
                "nome": venda.usuario.nome,
                "matricula": venda.usuario.matricula,
                "divida": venda.usuario.saldo_em_dividas,
                "total_compras": 0,
                "total_fiado": 0,
                "total_pago": 0,
            }

        resumo_clientes[uid]["total_compras"] += float(venda.valor_total)
        total_vendas += float(venda.valor_total)

        if venda.metodo_pagamento == "fiado":
            resumo_clientes[uid]["total_fiado"] += float(venda.valor_total)
            total_fiado += float(venda.valor_total)
        else:
            resumo_clientes[uid]["total_pago"] += float(venda.valor_total)
            total_pago += float(venda.valor_total)

    entradas = EntradaEstoque.objects.filter(data__date__gte=data_inicial)
    total_gasto_estoque = sum(float(e.valor_total()) for e in entradas)

    lucro_prejuizo = total_vendas - total_gasto_estoque

    context = {
        "periodo": periodo,
        "vendas": vendas,
        "resumo_clientes": list(resumo_clientes.values()),
        "entradas": entradas,
        "total_fiado": total_fiado,
        "total_pago": total_pago,
        "total_vendas": total_vendas,
        "total_gasto_estoque": total_gasto_estoque,
        "lucro_prejuizo": lucro_prejuizo,
    }

    return render(request, "Site_principal/Relatorio/gerar_relatorio.html", context)

from django.db.models import Sum, F
from datetime import datetime, timedelta
from .models import Usuario, Venda, ItemVenda, EntradaEstoque
from decimal import Decimal
from openpyxl import Workbook
from django.http import HttpResponse
from django.utils.timezone import localdate
import datetime

def exportar_relatorio(request):
    periodo = request.GET.get('periodo', 'hoje')

    agora = datetime.datetime.now()

    periodos = {
        "hoje": agora.replace(hour=0, minute=0, second=0, microsecond=0),
        "5d": agora - datetime.timedelta(days=5),
        "10d": agora - datetime.timedelta(days=10),
        "1m": agora - datetime.timedelta(days=30),
        "6m": agora - datetime.timedelta(days=180),
        "1a": agora - datetime.timedelta(days=365),
        "2a": agora - datetime.timedelta(days=365 * 2),
        "3a": agora - datetime.timedelta(days=365 * 3),
    }

    inicio = periodos.get(periodo, agora.replace(hour=0, minute=0, second=0, microsecond=0))

    # --------------------------- #
    # Resumo de clientes          #
    # --------------------------- #
    resumo_clientes = []
    usuarios = Usuario.objects.all()

    for u in usuarios:
        vendas_usuario = Venda.objects.filter(usuario=u, data__gte=inicio)

        total_compras = sum(Decimal(v.valor_total) for v in vendas_usuario)
        total_fiado = sum(Decimal(v.valor_total) for v in vendas_usuario if v.metodo_pagamento == 'fiado')
        total_pago = sum(Decimal(v.valor_total) for v in vendas_usuario if v.metodo_pagamento != 'fiado')

        resumo_clientes.append({
            'nome': u.nome,
            'matricula': u.matricula,
            'divida': Decimal(u.saldo_em_dividas),
            'total_compras': total_compras,
            'total_fiado': total_fiado,
            'total_pago': total_pago
        })

    # --------------------------- #
    # Vendas detalhadas           #
    # --------------------------- #
    vendas = Venda.objects.filter(data__gte=inicio).select_related("usuario", "vendedor")

    # --------------------------- #
    # Entradas de estoque         #
    # --------------------------- #
    entradas = EntradaEstoque.objects.filter(data__gte=inicio)

    # --------------------------- #
    # Criar Excel                 #
    # --------------------------- #
    wb = Workbook()

    # RESUMO CLIENTES
    ws1 = wb.active
    ws1.title = "Resumo Clientes"
    ws1.append(["Nome", "Matrícula", "Dívida (R$)", "Compras (R$)", "Fiado (R$)", "Pago (R$)"])

    for p in resumo_clientes:
        ws1.append([
            p['nome'],
            p['matricula'],
            float(p['divida']),
            float(p['total_compras']),
            float(p['total_fiado']),
            float(p['total_pago'])
        ])

    # VENDAS DETALHADAS
    ws2 = wb.create_sheet(title="Vendas Detalhadas")
    ws2.append(["Data", "Cliente", "Vendedor", "Total (R$)", "Pagamento"])

    for v in vendas:
        ws2.append([
            v.data.replace(tzinfo=None),
            v.usuario.nome,
            v.vendedor.nome if v.vendedor else "—",
            float(Decimal(v.valor_total)),
            v.metodo_pagamento
        ])

    # ENTRADAS ESTOQUE
    ws3 = wb.create_sheet(title="Entradas Estoque")
    ws3.append(["Data", "Produto", "Quantidade", "Preço Unitário (R$)", "Valor Total (R$)"])

    for e in entradas:
        valor_total = Decimal(e.quantidade) * Decimal(e.preco_unitario)
        ws3.append([
            e.data.replace(tzinfo=None),
            e.produto.nome,
            e.quantidade,
            float(Decimal(e.preco_unitario)),
            float(valor_total)
        ])

    # RESUMO GERAL
    ws4 = wb.create_sheet(title="Resumo Geral")
    ws4.append(["Descrição", "Valor (R$)"])

    total_fiado = sum(p['total_fiado'] for p in resumo_clientes)
    total_pago = sum(p['total_pago'] for p in resumo_clientes)
    total_vendas = sum(p['total_compras'] for p in resumo_clientes)
    total_gasto_estoque = sum(Decimal(e.quantidade) * Decimal(e.preco_unitario) for e in entradas)

    lucro_prejuizo = total_vendas - total_gasto_estoque

    ws4.append(["Total Fiado", float(total_fiado)])
    ws4.append(["Total Pago", float(total_pago)])
    ws4.append(["Total Vendas", float(total_vendas)])
    ws4.append(["Gasto no Estoque", float(total_gasto_estoque)])
    ws4.append(["Lucro / Prejuízo", float(lucro_prejuizo)])

    from openpyxl.utils import get_column_letter
    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    # --------------------------- #
    # Baixar arquivo              #
    # --------------------------- #
    data = localdate().strftime("%Y-%m-%d")
    nome_arquivo = f"relatorio_{data}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    wb.save(response)

    return response